"""
NEXUS MCP — Advanced Files Module
ZIP/TAR, CSV/JSON/Excel processing, PDF parse, Git local operations
"""
import json
import logging
import os
from typing import Optional
from pydantic import BaseModel, Field, ConfigDict
from mcp.server.fastmcp import FastMCP, Context

log = logging.getLogger("nexus-mcp.files_advanced")


def register(mcp: FastMCP):

    # ── ZIP / Archive ────────────────────────────────────

    class ZipCreateInput(BaseModel):
        model_config = ConfigDict(extra="forbid")
        output_path: str = Field(..., description="Output zip file path")
        source_paths: list = Field(..., description="List of files/dirs to zip")

    @mcp.tool(name="zip_create", annotations={"title": "Create ZIP Archive"})
    async def zip_create(params: ZipCreateInput, ctx: Context) -> str:
        """Create a ZIP archive from files/directories."""
        import zipfile
        try:
            with zipfile.ZipFile(params.output_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for src in params.source_paths:
                    if os.path.isdir(src):
                        for root, dirs, files in os.walk(src):
                            for file in files:
                                fp = os.path.join(root, file)
                                zf.write(fp, os.path.relpath(fp, os.path.dirname(src)))
                    else:
                        zf.write(src, os.path.basename(src))
            size = os.path.getsize(params.output_path)
            return json.dumps({"status": "ok", "path": params.output_path, "size_bytes": size})
        except Exception as e:
            return f"Error: {e}"

    class ZipExtractInput(BaseModel):
        model_config = ConfigDict(extra="forbid")
        zip_path: str = Field(..., description="ZIP file to extract")
        output_dir: str = Field(..., description="Directory to extract into")

    @mcp.tool(name="zip_extract", annotations={"title": "Extract ZIP Archive"})
    async def zip_extract(params: ZipExtractInput, ctx: Context) -> str:
        """Extract a ZIP archive to directory."""
        import zipfile
        try:
            os.makedirs(params.output_dir, exist_ok=True)
            with zipfile.ZipFile(params.zip_path, "r") as zf:
                zf.extractall(params.output_dir)
                names = zf.namelist()
            return json.dumps({"status": "ok", "extracted": len(names), "dir": params.output_dir, "files": names[:50]})
        except Exception as e:
            return f"Error: {e}"

    # ── CSV ───────────────────────────────────────────────

    class CSVReadInput(BaseModel):
        model_config = ConfigDict(extra="forbid")
        path: str = Field(..., description="CSV file path")
        limit: int = Field(100, description="Max rows to read")
        delimiter: str = Field(",", description="CSV delimiter")

    @mcp.tool(name="csv_read", annotations={"title": "Read CSV File"})
    async def csv_read(params: CSVReadInput, ctx: Context) -> str:
        """Read a CSV file and return rows as JSON."""
        import csv
        try:
            rows = []
            with open(params.path, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f, delimiter=params.delimiter)
                for i, row in enumerate(reader):
                    if i >= params.limit:
                        break
                    rows.append(dict(row))
            return json.dumps({"rows": rows, "count": len(rows), "path": params.path})
        except Exception as e:
            return f"Error: {e}"

    class CSVWriteInput(BaseModel):
        model_config = ConfigDict(extra="forbid")
        path: str = Field(..., description="Output CSV file path")
        rows: list = Field(..., description="List of dicts to write")
        delimiter: str = Field(",", description="CSV delimiter")

    @mcp.tool(name="csv_write", annotations={"title": "Write CSV File"})
    async def csv_write(params: CSVWriteInput, ctx: Context) -> str:
        """Write rows to a CSV file."""
        import csv
        try:
            if not params.rows:
                return json.dumps({"error": "No rows provided"})
            fieldnames = list(params.rows[0].keys())
            with open(params.path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=params.delimiter)
                writer.writeheader()
                writer.writerows(params.rows)
            return json.dumps({"status": "ok", "path": params.path, "rows": len(params.rows)})
        except Exception as e:
            return f"Error: {e}"

    # ── Excel ─────────────────────────────────────────────

    class ExcelReadInput(BaseModel):
        model_config = ConfigDict(extra="forbid")
        path: str = Field(..., description="Excel file path (.xlsx)")
        sheet: Optional[str] = Field(None, description="Sheet name (default: first sheet)")
        limit: int = Field(100, description="Max rows to read")

    @mcp.tool(name="excel_read", annotations={"title": "Read Excel File"})
    async def excel_read(params: ExcelReadInput, ctx: Context) -> str:
        """Read an Excel file and return rows as JSON."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(params.path, read_only=True, data_only=True)
            ws = wb[params.sheet] if params.sheet else wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return json.dumps({"rows": [], "count": 0})
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data = []
            for row in rows[1:params.limit + 1]:
                data.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
            wb.close()
            return json.dumps({"rows": data, "count": len(data), "sheet": ws.title, "columns": headers})
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"
        except Exception as e:
            return f"Error: {e}"

    class ExcelWriteInput(BaseModel):
        model_config = ConfigDict(extra="forbid")
        path: str = Field(..., description="Output Excel file path (.xlsx)")
        rows: list = Field(..., description="List of dicts to write")
        sheet_name: str = Field("Sheet1", description="Sheet name")

    @mcp.tool(name="excel_write", annotations={"title": "Write Excel File"})
    async def excel_write(params: ExcelWriteInput, ctx: Context) -> str:
        """Write rows to an Excel file."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = params.sheet_name
            if not params.rows:
                return json.dumps({"error": "No rows"})
            headers = list(params.rows[0].keys())
            ws.append(headers)
            for row in params.rows:
                ws.append([row.get(h, "") for h in headers])
            wb.save(params.path)
            return json.dumps({"status": "ok", "path": params.path, "rows": len(params.rows)})
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"
        except Exception as e:
            return f"Error: {e}"

    # ── PDF Parse ─────────────────────────────────────────

    class PDFReadInput(BaseModel):
        model_config = ConfigDict(extra="forbid")
        path: str = Field(..., description="PDF file path")
        pages: Optional[str] = Field(None, description="Page range, e.g. '1-5' or '3' (default: all, max 20)")

    @mcp.tool(name="pdf_read", annotations={"title": "Read/Parse PDF File"})
    async def pdf_read(params: PDFReadInput, ctx: Context) -> str:
        """Extract text from a PDF file."""
        try:
            import pypdf
            reader = pypdf.PdfReader(params.path)
            total = len(reader.pages)
            if params.pages:
                parts = params.pages.split("-")
                start = int(parts[0]) - 1
                end = int(parts[1]) if len(parts) > 1 else start + 1
            else:
                start, end = 0, min(total, 20)
            text = ""
            for i in range(start, min(end, total)):
                text += reader.pages[i].extract_text() + "\n\n"
            return json.dumps({"text": text[:20000], "total_pages": total, "extracted_pages": end - start})
        except ImportError:
            return "Error: pypdf not installed. Run: pip install pypdf"
        except Exception as e:
            return f"Error: {e}"

    # ── Git Local ─────────────────────────────────────────

    class GitInput(BaseModel):
        model_config = ConfigDict(extra="forbid")
        repo_path: str = Field(..., description="Local path to git repository")
        command: str = Field(..., description="Git subcommand: status, log, diff, add, commit, push, pull, clone, branch, checkout")
        args: Optional[str] = Field(None, description="Additional arguments as string")

    @mcp.tool(name="git_cmd", annotations={"title": "Git Local Command", "destructiveHint": True})
    async def git_cmd(params: GitInput, ctx: Context) -> str:
        """Run git commands on a local repository."""
        import asyncio
        cmd = f"git -C {params.repo_path} {params.command}"
        if params.args:
            cmd += f" {params.args}"
        try:
            proc = await asyncio.create_subprocess_shell(
                cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=60)
            return json.dumps({
                "exit_code": proc.returncode,
                "stdout": stdout.decode(errors="replace")[:8000],
                "stderr": stderr.decode(errors="replace")[:2000],
                "command": cmd,
            })
        except asyncio.TimeoutError:
            return json.dumps({"error": "Git command timed out"})
        except Exception as e:
            return f"Error: {e}"

    class GitCloneInput(BaseModel):
        model_config = ConfigDict(extra="forbid")
        repo_url: str = Field(..., description="Git repository URL")
        dest_path: str = Field(..., description="Local destination path")
        branch: Optional[str] = Field(None, description="Branch to checkout")
        depth: Optional[int] = Field(None, description="Shallow clone depth")

    @mcp.tool(name="git_clone", annotations={"title": "Git Clone Repository"})
    async def git_clone(params: GitCloneInput, ctx: Context) -> str:
        """Clone a git repository to local path."""
        import asyncio
        cmd = f"git clone {params.repo_url} {params.dest_path}"
        if params.branch:
            cmd += f" -b {params.branch}"
        if params.depth:
            cmd += f" --depth {params.depth}"
        try:
            proc = await asyncio.create_subprocess_shell(
                cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=120)
            return json.dumps({
                "exit_code": proc.returncode,
                "stdout": stdout.decode(errors="replace")[:5000],
                "stderr": stderr.decode(errors="replace")[:3000],
            })
        except asyncio.TimeoutError:
            return json.dumps({"error": "Clone timed out after 120s"})
        except Exception as e:
            return f"Error: {e}"

    log.info("Advanced Files module registered (ZIP, CSV, Excel, PDF, Git)")
